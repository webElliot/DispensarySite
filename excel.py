import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import os
from io import BytesIO

class Changes_Tracker:
	def __init__(self, online_data):
		self.keys_updated = 0
		self.rows_updated = 0

		self.History = {}
		self.load()
		self.online_data = online_data
		self.changed = self.verify_changes()



	def load(self):
		try:
			with open('previous/History.json' , 'r') as x:
				self.History = json.loads(x.read())
		except Exception as e:
			print(f"No history found.\nCreating Excel document.")

	def save_history(self):
		try:
			os.mkdir("previous")
			print(f"Created folder previous for storing History.")
			print("This is used to check whether entries have been updated.")
		except: pass
		try:
			with open('previous/History.json' , 'w') as x:
				x.write(json.dumps(self.online_data))
		except Exception as e:
			print(f"Failed to save History. : {e}")

	def build_dict(self,data_list):
		return {
			item['id'] : item
			for item in data_list
			}

	def compare_company(self, d1 , d2):
		keys_changed = 0
		for Key , Value in d1.items():
			D2_Value = d2.get(Key)
			if Value != D2_Value:
				keys_changed+=1
				#print(f"[{Value}] != [{D2_Value}]")
		return keys_changed
	def verify_changes(self):
		if not self.History:
			return True # No history.
		online = self.build_dict(json.loads(json.dumps(self.online_data)))
		offline = self.build_dict(self.History)
		#print(offline.keys())

		sorted_online = sorted(online.items())
		sorted_offline = sorted(offline.items())

		diff = 0
		same = 0
		for i , result in enumerate(sorted_online):
			if result!=sorted_offline[i]:
				for i2,k in enumerate(result):
					if k != sorted_offline[i][i2]:
						d1 = k
						d2 = sorted_offline[i][i2]

						keys_changed = self.compare_company(d1,d2)
						if keys_changed:
							diff += 1
							self.keys_updated += keys_changed
							self.rows_updated += 1
			else:
				same += 1

		print(f"{same} Companies had no changes.\n{diff} Companies updated their data for a total of {self.keys_updated} changes.")

		return 0 < diff # if difference is bigger than 0 there were changes.



class DownloadData:
	def __init__(self):
		self.session = requests.session()
		self.session.headers.update({
			"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
			})
		self.response = {}


	def request(self):
		try:
			r = requests.get('https://omma.us.thentiacloud.net/rest/public/profile/search/?keyword=all&skip=0&take=20&lang=en&type=Dispensary')
			if r.status_code == 200:
				result_count = r.json().get('resultCount')
				if result_count:
					print(f"Successfully downloaded {result_count} entries.")
				else:
					print(f"Page visited successfully but no products were returned.")
				self.response = r.json()
			else:
				print(f"Error {r.status_code} : {r.text}")

		except Exception as e:
			print(e)

	def run(self):
		self.request()


class Data_to_excel:
	COLUMN_WIDTHS = {
		'id': 30 ,
		'licenseNumber': 25 ,
		'legalName': 30 ,
		'tradeName': 30 ,
		'licenseType': 13 ,
		'streetAddress': 25 ,
		'city': 15 ,
		'county': 15 ,
		'licenseExpiryDate': 15 ,
		'zip': 15 ,
		'phone': 15 ,
		'email': 25 ,
		'hours': 25 ,
		'dataSourceName': 20 ,
		'discloseAddress': 15
		}
	
	Hide_Column = {
		"dataSourceName" : True
		}


	def __init__(self,data, byte_stream = False):
		self.byte_stream = byte_stream
		self.bytes = None
		self.document_created = False
		self.data = [Data_to_excel.hide_row(unique) for unique in data]
		self.run()
		
	@staticmethod
	def hide_row(data):
		for k in Data_to_excel.Hide_Column:
			if k in data:
				del data[k]
		return data
	
	def run(self):
		# Convert JSON data to a DataFrame
		df = pd.DataFrame(self.data)

		if self.byte_stream:
			output = BytesIO()

			# Save DataFrame to a BytesIO object
			with pd.ExcelWriter(output , engine = 'xlsxwriter') as writer:
				df.to_excel(writer , index = False)
			# writer.save() is not required here

			# Adjust column widths
			output.seek(0)  # Rewind the buffer
			self.adjust_column_widths(output)
			self.bytes = output
			return

		# Save DataFrame to an Excel file
		excel_file_path = 'licenses.xlsx'
		df.to_excel(excel_file_path , index = False)
		self.adjust_column_widths(excel_file_path)

		print("Successfully created excel document.")

	def adjust_column_widths (self , file_path):
		try:
			# Load the workbook
			wb = load_workbook(file_path)
			sheet = wb.active

			# Map DataFrame columns to Excel columns and set widths
			for index , column_name in enumerate(self.COLUMN_WIDTHS):
				column_letter = get_column_letter(index + 1)  # +1 because Excel columns start at 1, not 0
				sheet.column_dimensions[ column_letter ].width = self.COLUMN_WIDTHS[ column_name ]

			# Save the workbook
			wb.save(file_path)
			self.document_created = True
		except Exception as e:
			print(f"Failed to create document due to: {e}")

def create_doc():
	dataDownloader = DownloadData()
	dataDownloader.run()
	v = dataDownloader.response.get("result")

	if v:
		changes_Tracker = Changes_Tracker(v)

		if changes_Tracker.changed:
			print(f"Found changes!\nDocument pending creation.")
			excel_creator = Data_to_excel(v)
			if excel_creator.document_created:
				changes_Tracker.save_history()


if __name__ == "__main__":
	create_doc()

	""